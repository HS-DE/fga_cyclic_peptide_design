from __future__ import annotations

import argparse

from common import append_run_header, clean_sequence, load_config, resolve_path, setup_logger, write_fasta


def main() -> int:
    parser = argparse.ArgumentParser(description="从高丰度蛋白 Excel 中提取 FGA/P02671 序列。")
    parser.add_argument("--config", default="config/project.yaml")
    args = parser.parse_args()

    logger = setup_logger("01_extract_fga_sequence")
    append_run_header(logger, "01_extract_fga_sequence.py")
    config = load_config(args.config)

    # =====================
    # Step 1. 读取输入文件
    # =====================
    try:
        from openpyxl import Workbook, load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError("缺少 openpyxl，无法读取 Excel。请先安装 env/environment.yml。") from exc

    excel_path = resolve_path(config["input"]["excel_file"])
    if not excel_path.exists():
        raise FileNotFoundError(f"缺少输入 Excel: {excel_path}")
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Excel 为空: {excel_path}")
    header = [str(x).strip() if x is not None else "" for x in rows[0]]
    logger.info("Excel 列名: %s", header)

    gene_col = config["input"]["gene_column"]
    uniprot_col = config["input"]["uniprot_column"]
    seq_col = config["input"]["sequence_column"]
    missing_cols = [col for col in (gene_col, uniprot_col, seq_col) if col not in header]
    if missing_cols:
        raise ValueError(f"Excel 缺少列: {missing_cols}；请检查列名是否符合任务书。")
    gene_idx = header.index(gene_col)
    uniprot_idx = header.index(uniprot_col)
    seq_idx = header.index(seq_col)

    # =====================
    # Step 2. 提取 FGA/P02671 记录
    # =====================
    target_gene = config["project"]["target_gene"].upper()
    target_uniprot = config["project"]["target_uniprot"].upper()
    matches = []
    for row in rows[1:]:
        gene = str(row[gene_idx]).strip().upper() if row[gene_idx] is not None else ""
        uniprot = str(row[uniprot_idx]).strip().upper() if row[uniprot_idx] is not None else ""
        if gene == target_gene or uniprot == target_uniprot:
            matches.append(row)
    if not matches:
        raise ValueError("找不到 Gene == FGA 或 UniprotID == P02671 的记录；请检查列名和输入表。")
    selected = matches[0]
    sequence = clean_sequence(str(selected[seq_idx]))
    if not sequence:
        raise ValueError("FGA 记录存在，但 Sequence 为空。")
    if len(sequence) != 866:
        logger.warning("FGA 序列长度为 %s，不等于任务书预期 866；后续区域切片会按配置检查。", len(sequence))
    else:
        logger.info("FGA 序列长度确认: 866 aa")

    # =====================
    # Step 3. 输出 FGA 子表和 full-length FASTA
    # =====================
    out_xlsx = resolve_path("data/input/FGA_from_template.xlsx")
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "FGA"
    out_ws.append(header)
    for match in matches:
        out_ws.append(list(match))
    out_wb.save(out_xlsx)
    write_fasta("data/input/FGA_full_length_1_866.fasta", "FGA_P02671_full_length_1_866", sequence)
    logger.info("输出文件: %s", out_xlsx)
    logger.info("输出文件: data/input/FGA_full_length_1_866.fasta")
    logger.info("匹配记录数: %s", len(matches))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
